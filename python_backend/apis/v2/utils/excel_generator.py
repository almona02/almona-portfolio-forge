"""Excel generation utility for export functionality."""

from __future__ import annotations

import io
import json
from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore


def generate_excel_from_data(data: List[Dict[str, Any]]) -> bytes:
    """
    Generate Excel file (.xlsx) from list of dictionaries.

    Args:
        data: List of dictionaries (each dict represents a row)

    Returns:
        Excel file as bytes
    """
    if Workbook is None:
        raise ImportError(
            "openpyxl is required for Excel generation. "
            "Install it with: pip install openpyxl"
        )

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    if not data:
        # Return empty workbook with default sheet
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Extract headers from first row keys
    headers = list(data[0].keys())

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)

            # Format value
            if isinstance(value, (dict, list)):
                import json
                formatted_value = json.dumps(value, default=str)
            elif value is None:
                formatted_value = ""
            else:
                formatted_value = str(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
            cell.border = thin_border

            # Alternate row colors for readability
            if row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )

    # Auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        # Calculate max length (header + data)
        max_length = len(str(header))
        for row_data in data:
            value = row_data.get(header)
            if value is not None:
                value_str = str(value)
                if len(value_str) > 100:  # Limit to 100 chars for sizing
                    value_str = value_str[:100]
                max_length = max(max_length, len(value_str))
        # Set column width (min 10, max 50)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
